import json
import openpyxl
import re

def parse_path(path):
    """
    Convert a flattened path string into a list of keys/indices.
    Example: "a.b[0].c" -> ["a", "b", 0, "c"]
    """
    parts = re.split(r'\.(?![^\[]*\])', path)  # split on dots not inside []
    result = []
    for part in parts:
        # extract indices
        indices = re.findall(r'\[([0-9]+)\]', part)
        key = re.sub(r'\[[0-9]+\]', '', part)
        if key:
            result.append(key)
        for idx in indices:
            result.append(int(idx))
    return result

def set_value(target, path_parts, value):
    """
    Recursively set value in dict/list based on path parts.
    """
    current = target
    for i, p in enumerate(path_parts):
        is_last = i == len(path_parts) - 1
        if is_last:
            current[p] = value
        else:
            nxt = path_parts[i + 1]
            if isinstance(p, int):
                # Ensure list large enough
                while len(current) <= p:
                    current.append({} if not isinstance(nxt, int) else [])
                current = current[p]
            else:
                if p not in current:
                    current[p] = [] if isinstance(nxt, int) else {}
                current = current[p]

def excel_to_instance(excel_file, json_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Read header row and map header names to column indexes
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {name: idx for idx, name in enumerate(headers)}

    required_headers = ["path", "type", "value"]
    for h in required_headers:
        if h not in header_map:
            raise ValueError(f"Missing required header: '{h}' in Excel file")

    instance = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        path = row[header_map["path"]]
        type_ = row[header_map["type"]]
        value = row[header_map["value"]]

        if not path:
            continue

        # Restore JSON types
        if type_ == "array" or type_ == "dict":
            try:
                value = json.loads(value)
            except Exception:
                pass
        elif type_ == "int":
            try:
                value = int(value)
            except Exception:
                pass
        elif type_ == "float":
            try:
                value = float(value)
            except Exception:
                pass
        elif type_ == "bool":
            if isinstance(value, str):
                value = value.lower() in ("true", "1")

        path_parts = parse_path(path)
        set_value(instance, path_parts, value)

    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(instance, f, indent=2)

    print(f"JSON instance saved to {json_file}")


# Example usage
# excel_to_instance("../files/output/tests/FacilityRecord_instance_mapping_v2.xlsx", "../files/output/tests/FacilityRecord_instance_mapping_v2.json")